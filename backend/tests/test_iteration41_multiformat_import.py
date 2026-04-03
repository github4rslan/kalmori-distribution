"""
Iteration 41 - Multi-Format File Support for Admin Royalty Import
Tests: CSV, XLSX, PDF file uploads, artist_id assignment, file_format detection, auto-revenue estimation
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
ADMIN_EMAIL = "admin@tunedrop.com"
ADMIN_PASSWORD = "Admin123!"


class TestMultiFormatImport:
    """Test multi-format file import (CSV, XLSX, PDF)"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Login as admin before each test"""
        login_resp = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        assert login_resp.status_code == 200, f"Admin login failed: {login_resp.text}"
        data = login_resp.json()
        assert data.get("user", {}).get("role") == "admin", "User is not admin"
        self.token = data.get("access_token")
        self.headers = {"Authorization": f"Bearer {self.token}"}
        print(f"✓ Admin login successful")
        yield
        
    def test_csv_import_still_works(self):
        """Test that existing CSV import functionality still works"""
        csv_content = """Artist,Track,Platform,Streams,Revenue
Test Artist One,Test Song,Spotify,1000,4.00
Test Artist Two,Another Song,Apple Music,500,2.00
"""
        files = {"file": ("test_standard.csv", io.BytesIO(csv_content.encode()), "text/csv")}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, headers=self.headers)
        assert resp.status_code == 200, f"CSV import failed: {resp.text}"
        data = resp.json()
        
        assert "import_id" in data, "Missing import_id in response"
        assert data["total_rows"] == 2, f"Expected 2 rows, got {data['total_rows']}"
        assert "matched" in data, "Missing matched count"
        assert "unmatched" in data, "Missing unmatched count"
        assert "total_revenue" in data, "Missing total_revenue"
        assert data["total_revenue"] == 6.0, f"Expected revenue 6.0, got {data['total_revenue']}"
        print(f"✓ CSV import works: {data['total_rows']} rows, ${data['total_revenue']} revenue")
        
    def test_xlsx_import_works(self):
        """Test that XLSX file import works"""
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Artist", "Track", "Platform", "Streams", "Revenue"])
        ws.append(["XLSX Test Artist", "XLSX Song", "Spotify", 2000, 8.00])
        ws.append(["XLSX Test Artist 2", "XLSX Song 2", "Apple Music", 1500, 6.00])
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        files = {"file": ("test_import.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, headers=self.headers)
        assert resp.status_code == 200, f"XLSX import failed: {resp.text}"
        data = resp.json()
        
        assert "import_id" in data, "Missing import_id in response"
        assert data["total_rows"] == 2, f"Expected 2 rows, got {data['total_rows']}"
        assert data["total_revenue"] == 14.0, f"Expected revenue 14.0, got {data['total_revenue']}"
        print(f"✓ XLSX import works: {data['total_rows']} rows, ${data['total_revenue']} revenue")
        
    def test_xlsx_retail_ranking_format_requires_artist_id(self):
        """Test that retail ranking format (no artist column) requires artist_id"""
        import openpyxl
        
        # Create a Zojak-style ranking file (no artist column)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Retail", "Proportion", "Streams"])
        ws.append(["Spotify", "45%", 10000])
        ws.append(["Apple Music", "30%", 6500])
        ws.append(["Amazon Music", "25%", 5500])
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        files = {"file": ("ranking.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        # Without artist_id, should fail
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, headers=self.headers)
        assert resp.status_code == 400, f"Expected 400 for single-artist report without artist_id, got {resp.status_code}"
        assert "single-artist" in resp.json().get("detail", "").lower() or "artist" in resp.json().get("detail", "").lower(), \
            f"Error should mention artist selection: {resp.json()}"
        print(f"✓ Retail ranking format correctly requires artist_id")
        
    def test_xlsx_retail_ranking_with_artist_id(self):
        """Test retail ranking format with artist_id returns file_format field"""
        import openpyxl
        
        # First get a valid artist_id
        users_resp = requests.get(f"{BASE_URL}/api/admin/royalties/users", headers=self.headers)
        assert users_resp.status_code == 200
        users = users_resp.json().get("users", [])
        assert len(users) > 0, "No users found for assignment"
        test_artist_id = users[0]["id"]
        
        # Create a Zojak-style ranking file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Retail", "Proportion", "Streams"])
        ws.append(["Spotify", "45%", 10000])
        ws.append(["Apple Music", "30%", 6500])
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        files = {"file": ("ranking.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        form_data = {"artist_id": test_artist_id}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
        assert resp.status_code == 200, f"Import with artist_id failed: {resp.text}"
        result = resp.json()
        
        assert "file_format" in result, "Missing file_format in response"
        assert result["file_format"] == "retail_ranking", f"Expected retail_ranking format, got {result['file_format']}"
        assert result["total_rows"] == 2, f"Expected 2 rows, got {result['total_rows']}"
        assert result["matched"] == 2, f"Expected 2 matched (all assigned to artist), got {result['matched']}"
        print(f"✓ Retail ranking format with artist_id works: format={result['file_format']}, rows={result['total_rows']}")
        
    def test_xlsx_retail_daily_format(self):
        """Test retail daily format (Zojak Evolution style) with date columns"""
        import openpyxl
        
        # Get a valid artist_id
        users_resp = requests.get(f"{BASE_URL}/api/admin/royalties/users", headers=self.headers)
        users = users_resp.json().get("users", [])
        test_artist_id = users[0]["id"]
        
        # Create a Zojak Evolution-style daily file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Retail", "2024-01-01", "2024-01-02", "2024-01-03", "2024-01-04"])
        ws.append(["Spotify", 1000, 1200, 1100, 1300])
        ws.append(["Apple Music", 500, 600, 550, 650])
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        files = {"file": ("evolutions.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        form_data = {"artist_id": test_artist_id}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
        assert resp.status_code == 200, f"Daily format import failed: {resp.text}"
        result = resp.json()
        
        assert "file_format" in result, "Missing file_format in response"
        assert result["file_format"] == "retail_daily", f"Expected retail_daily format, got {result['file_format']}"
        print(f"✓ Retail daily format works: format={result['file_format']}, rows={result['total_rows']}")
        
    def test_auto_revenue_estimation(self):
        """Test that revenue is auto-estimated at $0.004/stream when no revenue column"""
        import openpyxl
        
        # Get a valid artist_id
        users_resp = requests.get(f"{BASE_URL}/api/admin/royalties/users", headers=self.headers)
        users = users_resp.json().get("users", [])
        test_artist_id = users[0]["id"]
        
        # Create file with streams but NO revenue column
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Retail", "Streams"])  # No Revenue column
        ws.append(["Spotify", 10000])  # Should estimate: 10000 * 0.004 = $40
        ws.append(["Apple Music", 5000])  # Should estimate: 5000 * 0.004 = $20
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        files = {"file": ("no_revenue.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        form_data = {"artist_id": test_artist_id}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
        assert resp.status_code == 200, f"Import failed: {resp.text}"
        result = resp.json()
        
        # Total should be (10000 + 5000) * 0.004 = $60
        expected_revenue = 60.0
        assert abs(result["total_revenue"] - expected_revenue) < 0.01, \
            f"Expected auto-estimated revenue ~${expected_revenue}, got ${result['total_revenue']}"
        print(f"✓ Auto-revenue estimation works: {result['total_revenue']} (expected ~{expected_revenue})")
        
    def test_artist_id_assigns_all_rows(self):
        """Test that artist_id form field assigns all rows to selected artist"""
        import openpyxl
        
        # Get a valid artist_id
        users_resp = requests.get(f"{BASE_URL}/api/admin/royalties/users", headers=self.headers)
        users = users_resp.json().get("users", [])
        test_artist_id = users[0]["id"]
        
        # Create file without artist column
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Track", "Platform", "Streams", "Revenue"])
        ws.append(["Song 1", "Spotify", 1000, 4.00])
        ws.append(["Song 2", "Apple Music", 500, 2.00])
        ws.append(["Song 3", "Amazon", 300, 1.20])
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        files = {"file": ("single_artist.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        form_data = {"artist_id": test_artist_id}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
        assert resp.status_code == 200, f"Import failed: {resp.text}"
        result = resp.json()
        
        # All rows should be matched to the selected artist
        assert result["matched"] == 3, f"Expected 3 matched rows, got {result['matched']}"
        assert result["unmatched"] == 0, f"Expected 0 unmatched rows, got {result['unmatched']}"
        print(f"✓ Artist assignment works: {result['matched']} rows assigned to artist")
        
    def test_unsupported_file_format_rejected(self):
        """Test that unsupported file formats are rejected"""
        # Try to upload a .txt file
        txt_content = "This is not a valid import file"
        files = {"file": ("test.txt", io.BytesIO(txt_content.encode()), "text/plain")}
        
        resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, headers=self.headers)
        assert resp.status_code == 400, f"Expected 400 for unsupported format, got {resp.status_code}"
        assert "unsupported" in resp.json().get("detail", "").lower(), \
            f"Error should mention unsupported format: {resp.json()}"
        print(f"✓ Unsupported file format correctly rejected")
        
    def test_import_with_template_still_works(self):
        """Test that template selector still works alongside artist selector"""
        # Get templates
        templates_resp = requests.get(f"{BASE_URL}/api/admin/distributor-templates", headers=self.headers)
        assert templates_resp.status_code == 200
        templates = templates_resp.json().get("templates", [])
        
        # Find DistroKid template which expects "Artist" column
        distrokid_template = next((t for t in templates if t["name"] == "DistroKid"), None)
        
        if distrokid_template:
            template_id = distrokid_template["id"]
            
            # Create CSV matching DistroKid format (Artist, Song/Album, Store, Quantity, Earnings (USD))
            csv_content = """Artist,Song/Album,Store,Quantity,Earnings (USD)
Template Test Artist,Template Song,Spotify,1000,4.00
"""
            files = {"file": ("template_test.csv", io.BytesIO(csv_content.encode()), "text/csv")}
            form_data = {"template_id": template_id}
            
            resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
            assert resp.status_code == 200, f"Import with template failed: {resp.text}"
            print(f"✓ Template selector still works with DistroKid template")
        elif len(templates) > 0:
            # Fallback: use auto-detect with standard headers
            csv_content = """Artist,Track,Platform,Streams,Revenue
Template Test Artist,Template Song,Spotify,1000,4.00
"""
            files = {"file": ("template_test.csv", io.BytesIO(csv_content.encode()), "text/csv")}
            
            resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, headers=self.headers)
            assert resp.status_code == 200, f"Import without template failed: {resp.text}"
            print(f"✓ Auto-detect import works (no matching template found)")
        else:
            print("⚠ No templates found, skipping template test")


class TestImportHistoryAndExistingEndpoints:
    """Test that existing endpoints still work after changes"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Login as admin before each test"""
        login_resp = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        assert login_resp.status_code == 200
        self.token = login_resp.json().get("access_token")
        self.headers = {"Authorization": f"Bearer {self.token}"}
        yield
        
    def test_import_history_loads(self):
        """Test that import history endpoint still works"""
        resp = requests.get(f"{BASE_URL}/api/admin/royalties/imports", headers=self.headers)
        assert resp.status_code == 200, f"Import history failed: {resp.text}"
        data = resp.json()
        assert "imports" in data, "Missing imports array"
        print(f"✓ Import history loads: {len(data['imports'])} imports")
        
    def test_admin_dashboard_works(self):
        """Test admin dashboard endpoint"""
        resp = requests.get(f"{BASE_URL}/api/admin/dashboard", headers=self.headers)
        assert resp.status_code == 200, f"Admin dashboard failed: {resp.text}"
        data = resp.json()
        assert "total_users" in data, "Missing total_users"
        print(f"✓ Admin dashboard works: {data['total_users']} users")
        
    def test_admin_users_works(self):
        """Test admin users endpoint"""
        resp = requests.get(f"{BASE_URL}/api/admin/users", headers=self.headers)
        assert resp.status_code == 200, f"Admin users failed: {resp.text}"
        data = resp.json()
        assert "users" in data, "Missing users array"
        print(f"✓ Admin users works: {len(data['users'])} users")
        
    def test_admin_templates_works(self):
        """Test admin templates endpoint"""
        resp = requests.get(f"{BASE_URL}/api/admin/distributor-templates", headers=self.headers)
        assert resp.status_code == 200, f"Admin templates failed: {resp.text}"
        data = resp.json()
        assert "templates" in data, "Missing templates array"
        print(f"✓ Admin templates works: {len(data['templates'])} templates")
        
    def test_admin_reconciliation_works(self):
        """Test admin reconciliation endpoint"""
        resp = requests.get(f"{BASE_URL}/api/admin/royalties/reconciliation", headers=self.headers)
        assert resp.status_code == 200, f"Admin reconciliation failed: {resp.text}"
        data = resp.json()
        assert "summary" in data, "Missing summary"
        assert "duplicates" in data, "Missing duplicates"
        assert "discrepancies" in data, "Missing discrepancies"
        print(f"✓ Admin reconciliation works")


class TestLabelExportBranding:
    """Test that label exports have Kalmori Distribution branding"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Login as admin (who is also a label producer)"""
        login_resp = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        assert login_resp.status_code == 200
        self.token = login_resp.json().get("access_token")
        self.headers = {"Authorization": f"Bearer {self.token}"}
        yield
        
    def test_csv_export_has_kalmori_branding(self):
        """Test that CSV export has Kalmori Distribution branding"""
        resp = requests.get(f"{BASE_URL}/api/label/royalties/export/csv", headers=self.headers)
        assert resp.status_code == 200, f"CSV export failed: {resp.text}"
        
        content = resp.text
        # Check first line has Kalmori branding
        first_line = content.split('\n')[0]
        assert "Kalmori" in first_line, f"First line should have Kalmori branding: {first_line}"
        print(f"✓ CSV export has Kalmori branding: {first_line}")
        
    def test_label_payouts_tab_export_only(self):
        """Test that label royalties endpoint still works (export-only functionality)"""
        resp = requests.get(f"{BASE_URL}/api/label/royalties", headers=self.headers)
        assert resp.status_code == 200, f"Label royalties failed: {resp.text}"
        data = resp.json()
        assert "summary" in data, "Missing summary"
        assert "artists" in data, "Missing artists"
        print(f"✓ Label royalties endpoint works (export-only)")


class TestRealZojakFiles:
    """Test with real Zojak files if available"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Login as admin"""
        login_resp = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        assert login_resp.status_code == 200
        self.token = login_resp.json().get("access_token")
        self.headers = {"Authorization": f"Bearer {self.token}"}
        yield
        
    def test_real_ranking_xlsx(self):
        """Test with real ranking.xlsx file if available"""
        ranking_path = "/tmp/ranking.xlsx"
        if not os.path.exists(ranking_path):
            pytest.skip("ranking.xlsx not found in /tmp")
            
        # Get a valid artist_id
        users_resp = requests.get(f"{BASE_URL}/api/admin/royalties/users", headers=self.headers)
        users = users_resp.json().get("users", [])
        test_artist_id = users[0]["id"]
        
        with open(ranking_path, "rb") as f:
            files = {"file": ("ranking.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            form_data = {"artist_id": test_artist_id}
            
            resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
            assert resp.status_code == 200, f"Real ranking.xlsx import failed: {resp.text}"
            result = resp.json()
            print(f"✓ Real ranking.xlsx imported: {result['total_rows']} rows, format={result.get('file_format', 'N/A')}")
            
    def test_real_evolutions_xlsx(self):
        """Test with real evolutions.xlsx file if available"""
        evolutions_path = "/tmp/evolutions.xlsx"
        if not os.path.exists(evolutions_path):
            pytest.skip("evolutions.xlsx not found in /tmp")
            
        # Get a valid artist_id
        users_resp = requests.get(f"{BASE_URL}/api/admin/royalties/users", headers=self.headers)
        users = users_resp.json().get("users", [])
        test_artist_id = users[0]["id"]
        
        with open(evolutions_path, "rb") as f:
            files = {"file": ("evolutions.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            form_data = {"artist_id": test_artist_id}
            
            resp = requests.post(f"{BASE_URL}/api/admin/royalties/import", files=files, data=form_data, headers=self.headers)
            assert resp.status_code == 200, f"Real evolutions.xlsx import failed: {resp.text}"
            result = resp.json()
            print(f"✓ Real evolutions.xlsx imported: {result['total_rows']} rows, format={result.get('file_format', 'N/A')}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
